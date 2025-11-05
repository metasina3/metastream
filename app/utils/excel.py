"""
Excel export utilities
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from fastapi.responses import Response
import tempfile
import os


def create_viewers_excel(viewers_data: list) -> Response:
    """
    Create Excel file for viewers data
    
    Returns: FastAPI Response with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Viewers"
    
    # Headers
    headers = ["ID", "نام", "شماره", "زمان ورود", "مدت تماشا (ثانیه)", "IP", "User Agent"]
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for viewer in viewers_data:
        ws.append([
            viewer.get("id"),
            viewer.get("name"),
            viewer.get("phone"),
            viewer.get("joined_at"),
            viewer.get("duration_watched"),
            viewer.get("ip_address", ""),
            viewer.get("user_agent", "")
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=viewers_{timestamp}.xlsx"}
    )


def export_viewers_to_excel(viewers, stream_id: int) -> str:
    """
    Generate an Excel file on disk for a given stream's viewers and return file path.
    This matches the expectation in routers.analytics.export_stream_viewers.
    """
    # Normalize ORM objects to dicts if needed
    rows = []
    for v in viewers:
        if isinstance(v, dict):
            rows.append(v)
        else:
            rows.append({
                "id": getattr(v, "id", None),
                "name": getattr(v, "name", None),
                "phone": getattr(v, "phone", None),
                "joined_at": getattr(v, "joined_at", None),
                "duration_watched": getattr(v, "duration_watched", None),
                "ip_address": getattr(v, "ip_address", ""),
                "user_agent": getattr(v, "user_agent", ""),
            })

    wb = Workbook()
    ws = wb.active
    ws.title = "Viewers"

    headers = ["ID", "نام", "شماره", "زمان ورود", "مدت تماشا (ثانیه)", "IP", "User Agent"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for viewer in rows:
        ws.append([
            viewer.get("id"),
            viewer.get("name"),
            viewer.get("phone"),
            getattr(viewer.get("joined_at"), "isoformat", lambda: viewer.get("joined_at"))(),
            viewer.get("duration_watched"),
            viewer.get("ip_address", ""),
            viewer.get("user_agent", ""),
        ])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Write to temp file and return path
    os.makedirs("/tmp", exist_ok=True)
    tmp = tempfile.NamedTemporaryFile(prefix=f"stream_{stream_id}_viewers_", suffix=".xlsx", delete=False, dir="/tmp")
    try:
        wb.save(tmp.name)
        tmp.flush()
    finally:
        tmp.close()
    return tmp.name
